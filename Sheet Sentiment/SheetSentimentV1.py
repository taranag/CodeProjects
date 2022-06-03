# This code takes in a sheet and returns a sentiment score for each sentence in a given column and row range
# Author: Taran Agnihotri
# Version: 1.0
# Last Updated: 3 June 2022

# Imports
import tkinter
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl as pyxl

# Sentiment Analysis Function
# text: the text to analyze
# returns: the sentiment of the text
def sentimentAnalysis(text):
    """Use a web service to determine whether the sentiment of text is positive""" 
    from urllib.request import urlopen
    from urllib.parse import urlencode
    from json import loads
    url = "http://text-processing.com/api/sentiment/"
    data = urlencode({"text": text}).encode("utf-8")
    print(data)
    response = urlopen(url, data)
    responseText = loads(response.read().decode("utf-8"))
    sentiment = responseText["label"]
    return sentiment

# Sentiment Analysis Function
# text: the text to analyze
# returns: the sentiment of the text, either positive or negative
def sentimentPosOrNeg(text):
    from urllib.request import urlopen
    from urllib.parse import urlencode
    from json import loads
    url = "http://text-processing.com/api/sentiment/"
    data = urlencode({"text": text}).encode("utf-8")
    print(data)
    response = urlopen(url, data)
    responseText = loads(response.read().decode("utf-8"))
    probabilities = responseText["probability"]
    if probabilities["neg"] > probabilities["pos"]:
        return "Negative, {0:.3f}".format(probabilities["neg"])
    else:
        return "Positive, {0:.3f}".format(probabilities["pos"])

# Cell Color Change Function
# Sets a set cell to a set color
# fileName: the file
def SheetColorChanger(fileName):
    '''changes color of a cell of a sheet using pyxl'''
    # open the file
    wb = pyxl.load_workbook(fileName)
    # get the sheet
    sheet = wb['Sheet1']
    # get the cell
    cell = sheet['A1']
    # change the color of the cell
    cell.fill = pyxl.styles.PatternFill(patternType='solid', fgColor='00FF00')
    # save the file
    wb.save(fileName)

#SheetColorChanger("testSheet.xlsx")

# Reads a sheet 
# fileName: the file
# row: the row to read
# col: the column to read
# returns: the text in the cell
def pyxlSheetReader(fileName, row, col):
    '''reads a cell of a sheet using pyxl'''
    
    # open the file
    wb = pyxl.load_workbook(fileName)
    # get the sheet
    sheet = wb['Sheet1']
    # get the cell
    cell = sheet[row][col]
    # return the cell
    return cell.value

# File selector using tkinter
# returns: the file path
def fileSelector(): 
    '''Use tkinter to create gui file selector''' 
    root = tkinter.Tk()
    root.withdraw()
    messagebox.showinfo("File Selector", "Choose file to analyze.")
    file_path = filedialog.askopenfilename()
    root.destroy()
    return file_path

# Data range selector using tkinter
# returns: the data range column, startRow, and endRow
def dataSelector():
    '''use tkinter to ask user for text input with column label''' 
    root = tkinter.Tk()
    root.withdraw()
    messagebox.showinfo("Data Selector", "Enter the column and row range to analyze.")
    colLabel = tk.Label(text="Column")
    colBox = tk.Entry()
    rowStartLabel = tk.Label(text="Start Row")
    rowStartBox = tk.Entry()
    rowEndLabel = tk.Label(text="End Row")
    rowEndBox = tk.Entry()
    confirmButton = tk.Button(text="Confirm", command=lambda: root.quit())

    colLabel.pack()
    colBox.pack()
    rowStartLabel.pack()
    rowStartBox.pack()
    rowEndLabel.pack()
    rowEndBox.pack()
    confirmButton.pack()
    
    root.deiconify()

    root.mainloop()
    # column = input("Enter column to analyze: ")
    # startRow = input("Enter start row: ")
    # endRow = input("Enter end row: ")
    return colBox.get(), rowStartBox.get(), rowEndBox.get()

# Main Function
# Asks user for file to analyze and data range to analyze
# Analyzes the data and places the sentiment of each sentence in a new column
def main():
    '''main function'''
    fileName = fileSelector()
    column, startRow, endRow = rowSelector()
    #fileName = "statements.xlsx"
    #column, startRow, endRow = "1", "1", "8"
    print("Filename: {}, Column: {}, Start Row: {}, End Row: {}".format(fileName, column, startRow, endRow))
    # Convert letter to number for column index
    if column.isalpha(): 
        column = column.upper()
        column = ord(column) - 64

    # Parse start and end row
    column = int(column)
    column = column - 1
    startRow = int(startRow)
    endRow = int(endRow)

    # Load workbook
    wb = pyxl.load_workbook(fileName)

    # Get the sheet
    sheet = wb['Sheet1']

    # Insert column after for sentiment analysis
    sheet.insert_cols(column+2)

    for i in range(startRow, endRow+1):
        # Get the cell
        cell = sheet[i][column]
        # Get the text
        text = cell.value
        # Get the sentiment
        sentiment = sentimentPosOrNeg(text)
        # Write the sentiment to the cell
        sheet[i][column+1].value = sentiment

    wb.save(fileName)
    
# Call main function
main()
#SheetColorChanger(fileSelector())
